import openpyxl
import pcbnew
import os
import wx
import re
import sys
import os.path
import os

from .version import __version__

class ExampleDialog(wx.Dialog):
    def __init__(self: "Import PCB datas from Excel file.", parent: wx.Frame) -> None:
        super().__init__(parent, -1, "Import PCB datas from Excel file.", size=(450, 200))

        panel = wx.Panel(self)        

        self.path = 'c:\\ProgramData\\PCBfromExcel\\'
        self.filename = 'PCBfromExcel.ini'
        if os.path.isdir(self.path) == False:
            os.mkdir(self.path)
        try:
            with open(self.path+self.filename) as txt_file:
                lines = [line.rstrip() for line in txt_file]
        except:
            lines = ['c:\\','4', '5']
            with open(self.path+self.filename, "w") as txt_file:
                for line in lines:
                    txt_file.write(line + "\n")

        l1 = wx.StaticText(panel,label = "File path:",style = wx.ALIGN_CENTRE, pos=(5,0)) 
        self.text_path = wx.TextCtrl(panel, 0, lines[0], pos=(10,20), size=(300,20))

        openBtn = wx.Button(panel, label='Open xlsx file', pos=(320,18))
        openBtn.Bind(wx.EVT_BUTTON, self.open_press)

        l1 = wx.StaticText(panel,label = "Those offsets is valid if in the xlsx file not by overwrite are!",style = wx.ALIGN_CENTRE, pos=(5, 50)) 

        l2 = wx.StaticText(panel,label = "Offset X:",style = wx.ALIGN_CENTRE, pos=(5, 65)) 
        self.text_x = wx.TextCtrl(panel, 0, lines[1], pos=(10, 85), size=(80,20))

        l3 = wx.StaticText(panel,label = "Offset Y:",style = wx.ALIGN_CENTRE, pos=(150, 65)) 
        self.text_y = wx.TextCtrl(panel, 0, lines[2], pos=(155, 85), size=(80, 20))

        startBtn = wx.Button(panel, label='Start process', pos=(50,120))
        startBtn.Bind(wx.EVT_BUTTON, self.start_press)

        cancelBtn = wx.Button(panel, label='Cancel', pos=(300,120))
        cancelBtn.Bind(wx.EVT_BUTTON, self.cancel_press)

        self.Centre()      
        self.Show()

    def start_press(self, event):
        #sys.excepthook = self.my_message
        try:
            lines = ['','', '']
            lines[0] = self.text_path.GetValue()
            lines[1] = self.text_x.GetValue()
            lines[2] = self.text_y.GetValue()
            with open(self.path+self.filename, "w") as txt_file:
                for line in lines:
                    txt_file.write(line + "\n")
        except:
            print('Error: The sysem can´t write the ini file.')
        self.SetCursor(wx.Cursor(wx.CURSOR_WAIT))
        offsetX  = float(self.text_x.GetValue())
        offsetY =  float(self.text_y.GetValue())
        pointNum = 0
        board = pcbnew.GetBoard()
        try:
            dataframe = openpyxl.load_workbook(self.text_path.GetValue())
            sheetNum =len(dataframe.worksheets)

            datas = pcbwriter(offsetX, offsetY, board)
            if (sheetNum>=1):
                dataframe1 = dataframe.worksheets[0]
                # Iterate the loop to read the cell values
                for rowNR in range(1, dataframe1.max_row+1):
                    datas.Clear();
                    for colNR in range(1, dataframe1.max_column+1):
                        MValue= dataframe1.cell(row =rowNR, column =colNR).value
                        self.DataSelecter(colNR, MValue, datas)
                    datas.dataProcessing(offsetX,offsetY,board)

            pcbnew.Refresh()
            print("Ok: "+str(datas.pointNum))    
            self.Destroy()
        except:
            self.SetCursor(wx.Cursor(wx.CURSOR_ARROW))
            dlg=wx.MessageDialog(None, 'The process went to error. Please check that the xlxs file is exist and this file content is all right.', 'Error',  wx.OK|wx.ICON_INFORMATION)
            dlg.ShowModal()
            dlg.Destroy()

    def cancel_press(self, event):
        self.Destroy()

    def open_press(self, event):
        value = self.text_path.GetValue()
        with wx.FileDialog(None, "Open xlsx file", wildcard="xlsx files (*.xlsx)|*.xlsx",
                                       style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return
            self.text_path.SetValue(fileDialog.GetPath())
        self.SetFocus()
        return

    def my_message(exception_type, exception_value, exception_traceback):
        msg = "Oh no! An error has occurred.\n\n"
        tb= traceback.format_exception(exception_type, exception_value, exception_traceback)
        for i in tb:
            msg += i
        dlg=wx.MessageDialog(None, msg, str(exception_type), wx.OK|wx.ICON_INFORMATION)
        dlg.ShowModal()
        dlg.Destroy()

    def DataSelecter(self, colNR, MValue, datas):
        if colNR == 1:
            datas.mpName=(str(MValue).strip(' '))
        if colNR == 2:
            datas.netName=(str(MValue).strip(' '))
        if colNR == 3:
            datas.pcbSide=(str(MValue).strip(' '))
        if colNR == 4:
            datas.headType=(str(MValue).strip(' '))
        if colNR == 5:
            datas.probeSize=datas.GetDiameter((str(MValue).strip(' ')))
        if colNR == 6:
            datas.x = -1
            datas.y = -1
            datas.posType = ''
            if str(MValue).find('a(')!=-1:
                datas.posType = 'absolute' # Tis position is from PCB absolute position
                a = str(MValue).replace("a(", "")
            else:
                a = str(MValue).replace("(", "")
            a = a.replace(")", "")
            a = a.replace(" ", "")
            location = a.split(",")
            if (len(location) == 2):
                datas.x=float(location[0])
                datas.y=float(location[1])
        if colNR == 7:
            datas.sch=(str(MValue).strip(' '))
        if colNR == 8:
            datas.footprint=(str(MValue).strip(' '))


class pcbwriter:

    mpName =""
    netName =""
    pcbSide =""
    headType =""
    probeSize = float(0)
    posType = ''
    x = float(0)
    y = float(0)
    net = pcbnew.NETINFO_ITEM
    board = pcbnew.BOARD
    offsetX = float(0)
    offsetY = float(0)
    offsetXxlsx = float(0)
    offsetYxlsx = float(0)
    x1 = float(0)
    y1 = float(0)
    x2 = float(0)
    y2 = float(0)
    pointNum = 0
    footprintPath = ''
    sch = ''
    footprint = ''

    def __init__(self, offsetX, offsetY, board):
        self.offsetX = offsetX
        self.offsetY = offsetY
        self.board = board

    def Clear(self):
        self.mpName =""
        self.netName =""
        self.pcbSide =""
        self.headType =""
        self.probeSize = float(0)
        self.x = float(0)
        self.y = float(0)
        footprintPath = pcbnew.GetWizardsSearchPaths()
        i0 = footprintPath.find('scripting\n')
        self.footprintPath = footprintPath [0:i0]+'footprints\\'

    def WriteTextToPCB(self, board):
        size = float(self.netName)
        pcb_txt = pcbnew.PCB_TEXT(self.board)
        pcb_txt.SetText(self.mpName)
        pcb_txt.SetPosition(pcbnew.wxPointMM(self.x, self.y))
        pcb_txt.SetHorizJustify(pcbnew.GR_TEXT_HJUSTIFY_CENTER)
        pcb_txt.Rotate(pcbnew.wxPointMM(self.x, self.y), self.probeSize*10)# 360fok = 3600
        pcb_txt.SetTextSize(pcbnew.wxSizeMM(size, size))
        if self.pcbSide == 'Top':
            pcb_txt.SetLayer(pcbnew.F_Cu)
        if self.pcbSide == 'Bottom':
            pcb_txt.SetLayer(pcbnew.B_Cu)
        if self.pcbSide == 'Top_Silks':
            pcb_txt.SetLayer(pcbnew.F_SilkS)
        if self.pcbSide == 'Bottom_Silks':
            pcb_txt.SetLayer(pcbnew.B_SilkS)
        board.Add(pcb_txt)

    def WritingaPCBBorder(self, board):
        rect = pcbnew.PCB_SHAPE(self.board)
        rect.SetShape(pcbnew.SHAPE_T_RECT)
        rect.SetFilled(False)
        rect.SetStart(pcbnew.wxPointMM(self.x1, self.y1))
        rect.SetEnd(pcbnew.wxPointMM(self.x2, self.y2))
        rect.SetLayer(pcbnew.Edge_Cuts)
        rect.SetWidth(int(0.1 * pcbnew.IU_PER_MM))
        board.Add(rect)

    def WriteMeassurePoint(self):
        # az adatok alapján az eljárás felrak egy mérőpontot a panelre
        self.net = self.board.FindNet(self.netName)
        if self.net is None:
            self.net = pcbnew.NETINFO_ITEM(self.board, self.netName)
            self.board.Add(self.net)
        module = self.board.FindFootprintByReference(self.mpName)
        if module is None:
            location = self.footprint.split(".pretty.")
            module = pcbnew.FootprintLoad(location[0]+".pretty",location[1])
            self.board.Add(module)
        module.SetPosition(pcbnew.wxPointMM(self.x, self.y))
        module.SetReference(self.mpName)
        module.SetLocked(True)
        newpad = module.Pads()[0]
        newpad.SetNetCode(self.net.GetNetCode())
        module.Reference().SetVisible(False)
        module.Value().SetVisible(False)

    def GetDiameter(self, diameterStr):
        s0 = ""
        for i0 in range(0, len(diameterStr)):
            if ((diameterStr[i0]>="0")and(diameterStr[i0]<="9")):
                s0 = s0 + diameterStr[i0]
            if ((diameterStr[i0]==".")or(diameterStr[i0]==",")or(diameterStr[i0]=="-")):
                s0 = s0 + diameterStr[i0]
        if s0 =='':
            s0 = '0'
        result = float(s0)
        isMil = diameterStr.find("mil")
        if not(isMil == -1):
            result = result * 0.0254
        return result

    def WriteHole(self):
        module = self.board.FindFootprintByReference(self.mpName)
        if module is None:
            location = self.footprint.split(".pretty.")
            module = pcbnew.FootprintLoad(location[0]+".pretty",location[1])
            self.board.Add(module)
        module.SetPosition(pcbnew.wxPointMM(self.x, self.y))
        module.SetReference(self.mpName)
        module.Reference().SetVisible(True)
        module.Value().SetVisible(False)

    def WritePart(self):
        module = self.board.FindFootprintByReference(self.mpName)
        if module is None:
            location = self.footprint.split(".pretty.")
            module = pcbnew.FootprintLoad(location[0]+".pretty",location[1])
            self.board.Add(module)
        module.SetPosition(pcbnew.wxPointMM(self.x, self.y))
        module.SetReference(self.mpName)
        module.Reference().SetVisible(True)
        module.Value().SetVisible(False)
        module.SetLocked(False)
        module.SetOrientation(self.probeSize*10)

    def WritePinNet(self):
        module = self.board.FindFootprintByReference(self.mpName)
        if not(module is None):
            pinNr = int(self.probeSize)
            #try:
            pin = module.FindPadByNumber(str(pinNr))
            moduleMP = self.board.FindFootprintByReference(self.netName)
            if not(moduleMP is None):
                #try:
                pinMp = moduleMP.Pads()[0]
                pin.SetNetCode(pinMp.GetNetCode())
            else:
                if not(self.netName=="NoNet"):
                    self.net = self.board.FindNet(self.netName)
                    if self.net is None:
                        print("Create a new network by name is "+self.netName+".")
                        self.net = pcbnew.NETINFO_ITEM(self.board, self.netName)
                        self.board.Add(self.net)
                pin.SetNetCode(self.net.GetNetCode())
        else:
            print("Error: Not found the target footprint. (Footprint ref: "+self.mpName+")")

    def dataProcessing(self, offsetX, offsetY, board):
        if (self.x == -1):
            if (self.headType == "Pin"):
                self.WritePinNet()
            if (self.headType == "OffsetX"):
                self.offsetXxlsx = self.probeSize
            if (self.headType == "OffsetY"):
                self.offsetYxlsx = self.probeSize
        else:
            if self.posType =='absolute':
                self.x = float(self.x)/1000
                self.y = float(self.y)/1000
            else:
                if (self.offsetXxlsx == 0)and(self.offsetYxlsx == 0):
                    self.x = offsetX-float(self.x)/1000
                    self.y = float(self.y)/1000-offsetY
                else:
                    self.x = self.offsetXxlsx-float(self.x)/1000
                    self.y = float(self.y)/1000-self.offsetYxlsx
            self.pointNum = self.pointNum + 1
            if (self.headType == "Spear")or(self.headType == "Crown")or(self.headType == "Flat"):
                self.WriteMeassurePoint()                    
            if (self.headType == "Corner1"):
                self.x1 = self.x
                self.y1 = self.y
            if (self.headType == "Corner2"):
                self.x2 = self.x
                self.y2 = self.y
                self.WritingaPCBBorder(board)
            if (self.headType == "Hole"):
                self.WriteHole()
            if (self.headType == "Part"):
                self.WritePart()
            if (self.headType == "Text"):
                self.WriteTextToPCB(board)
